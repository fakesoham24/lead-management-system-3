"""
Lead Routes — Full CRUD for leads, Excel upload, and OCR image scanning.
All endpoints are protected and scoped to the current user.
"""

import os
import io
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
import aiosqlite

from backend.database import get_db
from backend.models.lead import LeadCreate, LeadResponse, LeadUpdate
from backend.middleware.auth_middleware import get_current_user
from backend.services.lead_service import (
    create_lead,
    get_leads_for_user,
    get_lead_by_id,
    update_lead,
    delete_lead,
    delete_all_leads,
    get_lead_stats,
    ensure_user_upload_dir,
)

router = APIRouter(prefix="/api/leads", tags=["Leads"])


# ──────────────────────────────────────────────
#  Create a Lead Manually
# ──────────────────────────────────────────────
@router.post("", response_model=LeadResponse)
async def create_new_lead(
    lead_data: LeadCreate,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Create a new lead manually."""
    try:
        lead = await create_lead(
            db=db,
            owner_id=current_user["id"],
            company_name=lead_data.company_name,
            contact_name=lead_data.contact_name,
            email=lead_data.email,
            phone=lead_data.phone,
            source=lead_data.source,
            notes=lead_data.notes,
        )
        return lead
    except ValueError as e:
        if "Duplicate lead" in str(e):
            raise HTTPException(status_code=400, detail=str(e))
        raise e


# ──────────────────────────────────────────────
#  List All Leads (with optional filters)
# ──────────────────────────────────────────────
@router.get("")
async def list_leads(
    status_filter: str = Query(None, alias="status"),
    source: str = Query(None),
    search: str = Query(None),
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    List leads. Salesperson sees their own; Admin sees all.
    Optional filters: ?status=new&search=acme&source=Website
    """
    leads = await get_leads_for_user(
        db=db,
        user_id=current_user["id"],
        role=current_user["role"],
        status=status_filter,
        search=search,
        source=source,
    )
    return leads


# ──────────────────────────────────────────────
#  Get Lead Statistics (for dashboard)
# ──────────────────────────────────────────────
@router.get("/stats")
async def lead_stats(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get lead statistics for the dashboard."""
    stats = await get_lead_stats(db, current_user["id"], current_user["role"])
    return stats


# ──────────────────────────────────────────────
#  Delete All Leads
# ──────────────────────────────────────────────
@router.delete("/all")
async def delete_all_existing_leads(
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Delete all leads and their associated campaigns for the user."""
    count = await delete_all_leads(db, current_user["id"], current_user["role"])
    return {"message": f"Deleted {count} leads successfully"}


# ──────────────────────────────────────────────
#  Get a Single Lead
# ──────────────────────────────────────────────
@router.get("/{lead_id}", response_model=LeadResponse)
async def get_single_lead(
    lead_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Get a single lead by ID."""
    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    # Ensure user can only access their own leads (unless admin)
    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    return lead


# ──────────────────────────────────────────────
#  Update a Lead
# ──────────────────────────────────────────────
@router.put("/{lead_id}", response_model=LeadResponse)
async def update_existing_lead(
    lead_id: int,
    lead_data: LeadUpdate,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Update an existing lead."""
    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    updated = await update_lead(
        db=db,
        lead_id=lead_id,
        company_name=lead_data.company_name,
        contact_name=lead_data.contact_name,
        email=lead_data.email,
        phone=lead_data.phone,
        notes=lead_data.notes,
        status=lead_data.status,
    )
    return updated


# ──────────────────────────────────────────────
#  Delete a Lead
# ──────────────────────────────────────────────
@router.delete("/{lead_id}")
async def delete_existing_lead(
    lead_id: int,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Delete a lead and its associated campaigns."""
    lead = await get_lead_by_id(db, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    if current_user["role"] != "admin" and lead["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Access denied")

    await delete_lead(db, lead_id)
    return {"message": "Lead deleted successfully"}


# ──────────────────────────────────────────────
#  Upload Excel File (Bulk Import)
# ──────────────────────────────────────────────
@router.post("/upload/excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Upload an Excel file to bulk-import leads.
    
    Expected columns (case-insensitive):
      company_name | contact_name | email | phone | notes
    
    At least one of company_name/contact_name/email/phone must be present.
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Please upload an .xlsx or .xls file.",
        )

    try:
        import openpyxl

        # Save file temporarily to user's upload directory
        user_dir = ensure_user_upload_dir(current_user["id"])
        file_path = os.path.join(user_dir, file.filename)

        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        try:
            # Parse Excel file directly from memory (file_path only needed for OCR)
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active

            # Get headers from first row
            headers = []
            for cell in ws[1]:
                val = str(cell.value).strip().lower().replace(" ", "_") if cell.value else ""
                headers.append(val)

            # Map columns
            col_map = {}
            for i, h in enumerate(headers):
                if "company" in h:
                    col_map["company_name"] = i
                elif "contact" in h or "name" in h:
                    col_map["contact_name"] = i
                elif "email" in h or "mail" in h:
                    col_map["email"] = i
                elif "phone" in h or "mobile" in h or "tel" in h:
                    col_map["phone"] = i
                elif "note" in h or "remark" in h:
                    col_map["notes"] = i

            # Import rows
            imported = 0
            skipped = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    company = str(row[col_map["company_name"]]).strip() if "company_name" in col_map and row[col_map["company_name"]] else None
                    contact = str(row[col_map["contact_name"]]).strip() if "contact_name" in col_map and row[col_map["contact_name"]] else None
                    email = str(row[col_map["email"]]).strip() if "email" in col_map and row[col_map["email"]] else None
                    phone = str(row[col_map["phone"]]).strip() if "phone" in col_map and row[col_map["phone"]] else None
                    notes = str(row[col_map["notes"]]).strip() if "notes" in col_map and row[col_map["notes"]] else None

                    # Skip completely empty rows
                    if not any([company, contact, email, phone]):
                        skipped += 1
                        continue

                    # Clean up "None" strings
                    if company == "None": company = None
                    if contact == "None": contact = None
                    if email == "None": email = None
                    if phone == "None": phone = None
                    if notes == "None": notes = None

                    await create_lead(
                        db=db,
                        owner_id=current_user["id"],
                        company_name=company,
                        contact_name=contact,
                        email=email,
                        phone=phone,
                        source="excel",
                        notes=notes,
                    )
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
                    skipped += 1

            return {
                "message": f"Import complete: {imported} leads imported, {skipped} skipped",
                "imported": imported,
                "skipped": skipped,
                "errors": errors[:10],  # Return max 10 errors
                "file": file.filename,
            }

        finally:
            # Always delete the uploaded file after processing to prevent disk buildup
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception:
                pass  # Non-critical cleanup

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel processing failed: {str(e)}")



# ──────────────────────────────────────────────
#  Upload Visiting Card Image (OCR)
# ──────────────────────────────────────────────
@router.post("/upload/ocr")
async def upload_ocr(
    file: UploadFile = File(...),
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Upload a visiting card image and extract contact data using OCR.
    Supported formats: jpg, jpeg, png, bmp, tiff
    
    Returns the extracted data for review before saving.
    """
    # Validate file type
    allowed_extensions = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif')
    if not file.filename.lower().endswith(allowed_extensions):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Supported: {', '.join(allowed_extensions)}",
        )

    try:
        from backend.services.ocr_service import extract_text_from_image, parse_business_card

        # Save file temporarily to user's upload directory
        user_dir = ensure_user_upload_dir(current_user["id"])
        file_path = os.path.join(user_dir, f"card_{file.filename}")

        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        try:
            # Run OCR
            raw_text = extract_text_from_image(file_path)

            # Parse structured data from OCR text
            parsed = parse_business_card(raw_text)

            return {
                "message": "Card scanned successfully",
                "extracted": parsed,
                "file": file.filename,
            }

        finally:
            # Always delete the uploaded image after processing to prevent disk buildup
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception:
                pass  # Non-critical cleanup

    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")



# ──────────────────────────────────────────────
#  Save OCR Result as Lead
# ──────────────────────────────────────────────
@router.post("/ocr/save", response_model=LeadResponse)
async def save_ocr_lead(
    lead_data: LeadCreate,
    db: aiosqlite.Connection = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Save OCR-extracted data as a new lead.
    Called after the user reviews and edits the extracted data.
    """
    try:
        lead = await create_lead(
            db=db,
            owner_id=current_user["id"],
            company_name=lead_data.company_name,
            contact_name=lead_data.contact_name,
            email=lead_data.email,
            phone=lead_data.phone,
            source="ocr",
            notes=lead_data.notes,
        )
        return lead
    except ValueError as e:
        if "Duplicate lead" in str(e):
            raise HTTPException(status_code=400, detail=str(e))
        raise e
